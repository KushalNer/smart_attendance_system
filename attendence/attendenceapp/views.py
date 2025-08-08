from django.shortcuts import render, redirect ,get_object_or_404,HttpResponse
from django.contrib import messages
from .models import Teacher, Subject, Student, Allattendance
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
import os
import cv2
from datetime import datetime
import time
start_time = time.time()
import face_recognition
from collections import Counter
from openpyxl import Workbook #excel file download


import numpy as np
import pickle

# Create your views here.
def home(request):
    
    if request.user.is_authenticated:
        subjectdata = Subject.objects.filter(username = request.user)
        teacherdata = Teacher.objects.get(username = request.user)
        chart_data=[]
        barchart_data=[]

        now = datetime.now()
        datestore = now.strftime("%Y-%m-%d")
        for subject in subjectdata:
            data = Allattendance.objects.filter(asubject_name=subject.name, adate= datestore, ateacher_name =teacherdata.name)
            gender_counter = Counter([i.agender  for i in data])
            chart_data.append({
                'subject': subject.name,
                'labels': list(gender_counter.keys()),
                'values': list(gender_counter.values()),
                
            })
        #labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        #present_data = [25, 20, 22, 27, 26, 18, 0]
        #absent_data = [5, 10, 8, 3, 4, 12, 30]

        for subject in subjectdata:
            data = Allattendance.objects.filter(asubject_name=subject.name, adate= datestore, ateacher_name =teacherdata.name)
            student_data = Student.objects.filter(Class=subject.classname)
            no_of_student=student_data.count()
            present_data = data.count() 
            barchart_data.append({
                'labels':subject.name,
                'present_data': present_data,
                'absent_data':no_of_student-present_data,
            })

        return render(request,'home.html',{'chart_data': chart_data ,'subject_data':subjectdata,'barchart_data': barchart_data})
    else:
        return render(request,'index.html')
    

def barchart(request):
    if request.method == "POST":
        # teacherdata = Teacher.objects.get(username = request.user)
        # data = Allattendance.objects.filter(asubject_name=subject.name, ateacher_name =teacherdata.name)
        labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        data = [1, 0, 1, 1, 1, 0, 1]  # 1 = present, 0 = absent

        return render(request, 'home.html', {
            'labels': labels,
            'data': data,
        })
    



def register_teacher(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        contact = request.POST.get('phone')
        username= request.POST.get('username')
        password = request.POST.get('password')
        cpassword = request.POST.get('cpassword')
        image = request.FILES['image']

        if password == cpassword:
            if User.objects.filter(username=username).exists():
                messages.error(request, "This username already exists.")
                return render(request, "teacher_register.html")
            else:
                data = Teacher(name=name, email=email, contact= contact ,username = username,image=image, password=password)
                data.save()
                myuser = User.objects.create_user(username = username, email=email, password = password)
                myuser.save()
                messages.success(request,'Account Regsiter Successfully')
                return redirect('home')
        else:
            messages.error(request, "Password are Not Match")
            return render(request,"teacher_register.html")


    return render(request, 'teacher_register.html')


def user_login(request):
    if request.method == 'POST':
        name = request.POST.get('username')
        password = request.POST.get('password')

        user=authenticate(request, username = name, password = password)
        if user is not None:
            login(request,user)
            messages.success(request,f"{user.username} Login Sucessfully")
            return redirect('userprofile',request.user.username)
        else:
            messages.error(request,'Invalid Name and Password ')
            return redirect('loginaccount')
    return render(request,'login.html')


def user_logout(request):
    logout(request)
    messages.success(request,"Logout Successfully")
    return redirect('loginaccount')


def user_profile(request , username):
    user= get_object_or_404(User, username=username)
    user_data = Teacher.objects.get(username= user)
    subject_data = Subject.objects.filter(username=user)
    return render(request, "user_profile.html",{'user_data':user_data, 'subject_data':subject_data})


def subject(request,username):
    if request.method == 'POST':
        name = request.POST.get('name')
        user = username
        classname = request.POST.get('class')
        subject_data = get_object_or_404(Subject,username = user, name=name, classname = classname )

        if name not in subject_data.name:
            data= Subject(name = name, username = user, classname = classname)
            data.save()
            messages.success(request,"Subject Name Add Successfully")
        else:
            messages.success(request,"Subject Name already exist")

    return redirect('userprofile',user)


def delete_subject(request, subject_name):
    subject = get_object_or_404(Subject, name=subject_name, username=request.user)
    subject.delete()
    user=request.user
    return redirect('home')


def user_attendence(request):
    # for show subject name into get attendence system
    subject_data = Subject.objects.filter(username=request.user)
    sname = request.GET.get('sname') 

    return render(request,"attendence.html",{'subject_data':subject_data })
   

    
#attendence data for attendence page show
subject_name = None #gloabal variable declare to use tow functions
attendence_date = None

def attendence_data(request):
    #if request.method == "GET":
    global subject_name
    global attendence_date
    subject_name = request.GET.get('subjectname')    
    attendence_date = request.GET.get('adate')
    teacherdata = Teacher.objects.get(username = request.user)#for the teacher name to check attendance only
  
    messages.success(request,f"Subject name : {subject_name} and date : {attendence_date}")
    data = Allattendance.objects.filter(asubject_name = subject_name , adate = attendence_date ,ateacher_name = teacherdata.name)
        
    return render(request,'attendence.html',{'attendence_data':data})

def datewise_download(request):
        global subject_name
        global attendence_date
   
        subject = subject_name
        attendence_d = attendence_date
        teacherdata = Teacher.objects.get(username = request.user)#for the teacher name to check attendance only

        print("Subject Name:", subject_name)
        print("Attendance Date:", attendence_date)
        print("Teacher Name:", teacherdata.name)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance Report on {attendence_date}"
        ws.append(["      "," ",f"Attendance Daily Report of {subject}"])
        ws.append([f"Date:{attendence_d}"])
        ws.append(["Roll No.", "Student Name", "Gender", "Class", "Date", "Time", "Subject", "Teacher"])
        data = Allattendance.objects.filter(asubject_name = subject , adate = attendence_d ,ateacher_name = teacherdata.name)
        print("Query Result Count:", data.count())
        for d in data:
                ws.append([
                    d.arno,
                    d.aname,
                    d.agender,
                    d.aclass,
                    d.adate,
                    d.atime,
                    d.asubject_name,
                    d.ateacher_name
                ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Attendance_Report_{subject}_{attendence_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


        
results = [] #global varibale declaration
analysis_info = [] # this global variable that store analysis info ex.start date etc
subjectname = None 
def data_analysis(request):
    subjects = Subject.objects.filter(username = request.user)
    global subjectname
    global results 
    results.clear()# means already data are remove
    global analysis_info
    

    subjectname = request.GET.get('subject')
    start_date = request.GET.get('startdate')
    end_date = request.GET.get('enddate')
    analysis_info.append({
        'subject_name':subjectname,
        'start_date':start_date,
        'end_date':end_date,
    })
    selectsubjectobject = Subject.objects.filter(name = subjectname, username = request.user).first()
    if selectsubjectobject:
        subjectclass = selectsubjectobject.classname
        print(selectsubjectobject)
        print(subjectclass)
    else:
        print("subject not found")
    if start_date and end_date: 
        
        students = Student.objects.all().order_by('rno')
        teacherdata = Teacher.objects.get(username = request.user)

        #messages.success(request,f"Subject Name : {subjectname} || Start Date : {start_date} || End Date : {end_date} ")
        
        for student in students:
            if student.Class == subjectclass:
                total_lecture = Allattendance.objects.filter(asubject_name = subjectname, adate__range=(start_date, end_date), ateacher_name = teacherdata.name).values('adate').distinct().count()
                attend_lecture = Allattendance.objects.filter(aname=student.name, asubject_name = subjectname, adate__range=(start_date, end_date), ateacher_name = teacherdata.name ).count()

                percentage= round((attend_lecture / total_lecture)*100, 2)if total_lecture > 0 else 0
                results.append({
                    'rno':student.rno,
                    'sname':student.name,
                    'gender':student.gender,
                    'tlecture':total_lecture,
                    'alecture':attend_lecture,
                    'percentage':percentage,
                })
    return render(request,"analysis.html",{'subjects':subjects,'subjectname':subjectname, 'results':results})

#excel file download
def analysis_report_downlaod(request):
    global results
    global subjectname
    global analysis_info
    wb =Workbook()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance Analysis Report of {subjectname}"
    ws.append(["      ",f"Attendance Analysis Report of {subjectname}"])
    for info in analysis_info:
        a=1
        if a == 1 and info['start_date']!= None:
            ws.append([f"Start Date:{info['start_date']}"])
            ws.append([f"End Date:{info['end_date']}"])
            break
        a=a+1
    ws.append(["Roll No.", "Student Name", "Gender", "Total Lecture", "Attend Lecture","Percentage"])
    
    for result in results:
        ws.append([
            result['rno'],
            result['sname'],
            result['gender'],
            result['tlecture'],
            result['alecture'],
            result['percentage'],

        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Attendance_Report_{subjectname}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    results.clear()
    return response
        


def contact_us(request):
    
    return render(request,"contactus.html")


# there are two functions for regsiter a face and recognize face
# student register function
def register_face(request):
     classdata = Subject.objects.all().values('classname').distinct()# for class are display in a form
     if request.method == 'POST':
        name = request.POST.get('name').strip()
        email = request.POST['email']
        Class = request.POST['class']
        rno = request.POST['rno']
        gender = request.POST['gender']
        contact = request.POST['phone']
        
        #all data are save in a database
        data = Student(name = name, email = email, Class = Class, rno = rno, contact = contact, gender=gender)
        data.save()

        dataset_path = os.path.join('dataset', name)
        os.makedirs(dataset_path, exist_ok=True)

        cap = cv2.VideoCapture(0)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

        count = 0
        max_images = 10

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                count += 1
                face = frame[y:y+h, x:x+w]
                face = cv2.resize(face, (200, 200))

                file_path = os.path.join(dataset_path, f"{count}.jpg")
                cv2.imwrite(file_path, face)

                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                cv2.putText(frame, f"{count}", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255,255,255), 2)

            cv2.imshow("Registering Faces - Press 'q' to stop", frame)

            if cv2.waitKey(1) & 0xFF == ord('q') or count >= max_images:
                break

        cap.release()
        cv2.destroyAllWindows()
        #return HttpResponse(f"Face images for {name} saved successfully.")
        messages.success(request,f" {name} Register successfully.")


        # this part code for encode images for  register student 
        target_student = name # only register student encode faces not other student
        dataset_path = os.path.join(os.getcwd(), 'dataset')
        encoding_file = os.path.join(os.getcwd(), 'encodings.pkl')
        student_folder = os.path.join(dataset_path, target_student)

        known_encodings = []
        known_names = []

         # 🔁 Load existing encodings if available
        if os.path.exists(encoding_file):
            with open(encoding_file, 'rb') as f:
                known_encodings, known_names = pickle.load(f)

        # ✅ Check if student folder exists
        if not os.path.isdir(student_folder):
            print(f"❌ Student folder not found: {student_folder}")
            return

        print(f"📂 Found folder for: {target_student}")

        # 🔍 Loop through images and encode
        for img_file in os.listdir(student_folder):
            img_path = os.path.join(student_folder, img_file)
            print(f"🖼️ Loading image: {img_path}")
            try:
                img = face_recognition.load_image_file(img_path)
                encodings = face_recognition.face_encodings(img)
                if encodings:
                    known_encodings.append(encodings[0])
                    known_names.append(target_student)
                    print(f"✅ Encoded: {target_student}")
                else:
                    print(f"⚠️ No face found in: {img_path}")
            except Exception as e:
                print(f"❌ Error with {img_path}: {e}")


        print("Known Students:", known_names) 
        # 💾 Save updated encodings
        with open(encoding_file, 'wb') as f:
            pickle.dump((known_encodings, known_names), f)
        print("✅ Updated encodings saved.")

     return render(request, 'student_register.html',{'classdata':classdata})





def recognize(request, username):
    dataset_path = os.path.join(os.getcwd(), 'dataset')
    encoding_file = os.path.join(os.getcwd(), 'encodings.pkl')

    known_encodings = []
    known_names = []

    # ✅ Load precomputed encodings from pickle
    if os.path.exists(encoding_file):
        with open(encoding_file, 'rb') as f:
            known_encodings, known_names = pickle.load(f)
        #print("Known Students:", known_names) 
        
    else:
    
        for student_name in os.listdir(dataset_path):
            student_folder = os.path.join(dataset_path, student_name)
            if not os.path.isdir(student_folder):
                #print(f"❌ Skipping: {student_folder} is not a folder")
                continue

        # print(f"📂 Found folder for: {student_name}")

            for img_file in os.listdir(student_folder):
                    img_path = os.path.join(student_folder, img_file)
                #print(f"🖼️ Loading image: {img_path}")
                # try:
                    img = face_recognition.load_image_file(img_path)
                    encodings = face_recognition.face_encodings(img)
                    if encodings:
                        known_encodings.append(encodings[0])
                        known_names.append(student_name)
                        #print(f"✅ Encoded: {student_name}")
                    # else:
                    #     print(f"⚠️ No face found in: {img_path}")
                # except Exception as e:
                #     print(f"❌ Error with {img_path}: {e}")

        #print("Known Students:", known_names) 

    # 🎥 Open and configure webcam
    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

    if not cap.isOpened():
        messages.error(request, "❌ Cannot access webcam.")
        return render(request, 'home.html')

    marked = []
    start_time = time.time()
    unknown_alerted = False

    def mark_attendance(name):
        now = datetime.now()
        try:
            teacher = Teacher.objects.get(username=username)
            student = Student.objects.get(name=name)
            sname = request.GET.get('sname')
            datestore = now.strftime("%Y-%m-%d")

            if Allattendance.objects.filter(aname=name, asubject_name=sname, adate=datestore).exists():
                messages.success(request,f"⚠️ Already marked today for {name}")
                return

            Allattendance.objects.create(
                aname=name,
                aclass=student.Class,
                agender=student.gender,
                arno=student.rno,
                adate=datestore,
                atime=now.strftime('%H:%M:%S'),
                ateacher_name=teacher.name,
                asubject_name=sname
            )
            print(f"✅ Attendance marked for {name}")
            messages.success(request, f"✅ {name}, your attendance has been marked.")
        except Exception as e:
            print(f"❌ Error: {e}")
            messages.error(request, "Error while marking attendance.")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        for encoding, face_loc in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(known_encodings, encoding)
            distances = face_recognition.face_distance(known_encodings, encoding)

            if len(distances) == 0:
                continue

            best_match_index = np.argmin(distances)
            if matches[best_match_index]:
                name = known_names[best_match_index]
                if name not in marked:
                    mark_attendance(name)
                    marked.append(name)

                # Draw box
                top, right, bottom, left = [v * 4 for v in face_loc]
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.putText(frame, name, (left, top - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
            else:
                if not unknown_alerted:
                    print("⚠️ Unknown face detected.")
                    messages.warning(request, "⚠️ Unknown face detected.")
                    unknown_alerted = True

        cv2.imshow("Face Recognition Attendance", frame)

        # Break after 10 seconds or on 'q'
        if cv2.waitKey(1) & 0xFF == ord('q') or time.time() - start_time > 7:
            break

    cap.release()
    cv2.destroyAllWindows()
    return render(request, "attendence.html")




#this function are not be use
def recognizetp(request, username):
    dataset_path = os.path.join(os.getcwd(), 'dataset')
    known_encodings = []
    known_names = []

    # Load known images and encodings
    for student_name in os.listdir(dataset_path):
        student_folder = os.path.join(dataset_path, student_name)
        if not os.path.isdir(student_folder):
            continue
        for img_file in os.listdir(student_folder):
            img_path = os.path.join(student_folder, img_file)
            try:
                img = face_recognition.load_image_file(img_path)
                encodings = face_recognition.face_encodings(img)
                if encodings:
                    known_encodings.append(encodings[0])
                    known_names.append(student_name)
            except Exception as e:
                print(f"Error loading image {img_path}: {e}")
                continue

    # Start webcam
    cap = cv2.VideoCapture(0)
  
    if not cap.isOpened():
        messages.error(request, "❌ Webcam not accessible.")
        return render(request, 'home.html')

    marked = []
    start_time = time.time()

    def mark_attendance(name):
        try:
            now = datetime.now()
            teacher = Teacher.objects.get(username=username)
            student = Student.objects.get(name=name)
            sname = request.GET['sname'] 
            datestore = now.strftime("%Y-%m-%d")
            already = Allattendance.objects.filter(aname=name, asubject_name=sname, adate=datestore )
            if already.exists():
                messages.success(request,f"⚠️ Already marked today for {name}")
                return redirect('home')

            Allattendance.objects.create(
                aname=name,
                aclass=student.Class,
                agender=student.gender,
                arno=student.rno,
                adate = datestore,      
                atime= now.strftime('%H:%M:%S'),
                ateacher_name=teacher.name,  # Add this if you are storing as CharField
                asubject_name =sname
                # Add `asubject_name` if needed
            )
            messages.success(request, f"✅ {name}, thank you! Your attendance has been marked.")
            print(f"✅ Marked attendance for {name}")

            #attendence data show in a table format
            #attendence_data = Allattendance.objects.filter(asubject_name=sname, adate=datetime.now().date()) # all data for subject and date

        except Teacher.DoesNotExist:
            messages.error(request, "Teacher not found.")
        except Student.DoesNotExist:
            messages.error(request, f"Student {name} not found.")
        
        #return render(request,"attendence.html",{'attendence_data': attendence_data })

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # Resize for faster processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        # Detect faces
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        for encoding, face_loc in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(known_encodings, encoding)
            distances = face_recognition.face_distance(known_encodings, encoding)
            best_match_index = distances.argmin()

            if matches[best_match_index]:
                name = known_names[best_match_index]

                if name not in marked:
                    mark_attendance(name)
                    marked.append(name)

                # Scale back face location to original size
                top, right, bottom, left = [v * 4 for v in face_loc]
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                cv2.putText(frame, f"{name}", (left, top - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
            else:
                top, right, bottom, left = [v * 4 for v in face_loc]
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                cv2.putText(frame, "Unknown", (left, top - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                if not unknown_alerted:
                    print("⚠️ Unknown face detected.")
                    messages.warning(request, "⚠️ Unknown face detected.")
                    unknown_alerted = True
                
    
         
        # Display frame
        cv2.imshow("Face Recognition Attendance", frame)

        # Exit on 'q' or after 10 seconds
        if cv2.waitKey(1) & 0xFF == ord('q') or time.time() - start_time > 5:
            break

    cap.release()
    cv2.destroyAllWindows()
    #messages.success(request, "✅ Attendance process completed.")
    return render(request, "attendence.html")
